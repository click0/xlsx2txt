"""Excel file reader for xlsx2txt."""

from pathlib import Path
from typing import Any, Optional, Union
from datetime import datetime

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell as OpenpyxlCell
from openpyxl.utils import get_column_letter

from xlsx2txt.models import (
    Cell,
    CellStyle,
    FontStyle,
    FillStyle,
    BorderStyle,
    BorderSide,
    AlignmentStyle,
    Sheet,
    SheetDimensions,
    ColumnDimension,
    RowDimension,
    CellData,
)


def _get_color_value(color) -> Optional[str]:
    """Extract color value from openpyxl color object."""
    if color is None:
        return None
    if color.type == "rgb" and color.rgb:
        return color.rgb
    if color.type == "theme":
        return f"theme:{color.theme}"
    if color.type == "indexed":
        return f"indexed:{color.indexed}"
    return None


def _extract_font_style(cell: OpenpyxlCell) -> FontStyle:
    """Extract font style from openpyxl cell."""
    font = cell.font
    return FontStyle(
        name=font.name or "Calibri",
        size=font.size or 11.0,
        bold=font.bold or False,
        italic=font.italic or False,
        underline=font.underline,
        strike=font.strike or False,
        color=_get_color_value(font.color),
    )


def _extract_fill_style(cell: OpenpyxlCell) -> FillStyle:
    """Extract fill style from openpyxl cell."""
    fill = cell.fill
    return FillStyle(
        pattern_type=fill.patternType,
        fg_color=_get_color_value(fill.fgColor) if fill.fgColor else None,
        bg_color=_get_color_value(fill.bgColor) if fill.bgColor else None,
    )


def _extract_border_side(side) -> Optional[BorderSide]:
    """Extract single border side."""
    if side is None or side.style is None:
        return None
    return BorderSide(
        style=side.style,
        color=_get_color_value(side.color),
    )


def _extract_border_style(cell: OpenpyxlCell) -> BorderStyle:
    """Extract border style from openpyxl cell."""
    border = cell.border
    return BorderStyle(
        left=_extract_border_side(border.left),
        right=_extract_border_side(border.right),
        top=_extract_border_side(border.top),
        bottom=_extract_border_side(border.bottom),
    )


def _extract_alignment_style(cell: OpenpyxlCell) -> AlignmentStyle:
    """Extract alignment style from openpyxl cell."""
    alignment = cell.alignment
    return AlignmentStyle(
        horizontal=alignment.horizontal,
        vertical=alignment.vertical,
        wrap_text=alignment.wrap_text or False,
        text_rotation=alignment.text_rotation or 0,
    )


def _extract_cell_style(cell: OpenpyxlCell) -> CellStyle:
    """Extract complete style from openpyxl cell."""
    return CellStyle(
        font=_extract_font_style(cell),
        fill=_extract_fill_style(cell),
        border=_extract_border_style(cell),
        alignment=_extract_alignment_style(cell),
        number_format=cell.number_format or "General",
    )


def read_cell(
    path: Union[str, Path],
    coordinate: str = "A1",
    sheet_name: Optional[str] = None,
) -> Cell:
    """
    Read a cell from an Excel file.

    Args:
        path: Path to the Excel file.
        coordinate: Cell coordinate (e.g., "A1").
        sheet_name: Sheet name. If None, uses active sheet.

    Returns:
        Cell object with value and style.

    Raises:
        FileNotFoundError: If file does not exist.
        ValueError: If sheet_name is invalid.
    """
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"File not found: {path}")

    wb = load_workbook(path, data_only=False)

    if sheet_name:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet not found: {sheet_name}")
        sheet = wb[sheet_name]
    else:
        sheet = wb.active

    cell = sheet[coordinate]

    # Determine data type
    if cell.value is None:
        data_type = "n"
    elif isinstance(cell.value, bool):
        data_type = "b"
    elif isinstance(cell.value, (int, float)):
        data_type = "n"
    elif isinstance(cell.value, str):
        data_type = "s"
    else:
        data_type = "s"

    # Get formula if exists
    formula = None
    if cell.data_type == "f" or (isinstance(cell.value, str) and cell.value.startswith("=")):
        formula = cell.value

    return Cell(
        coordinate=coordinate,
        value=cell.value,
        data_type=data_type,
        formula=formula,
        style=_extract_cell_style(cell),
    )


def read_cell_style(
    path: Union[str, Path],
    coordinate: str = "A1",
    sheet_name: Optional[str] = None,
) -> CellStyle:
    """
    Read only the style of a cell from an Excel file.

    Args:
        path: Path to the Excel file.
        coordinate: Cell coordinate (e.g., "A1").
        sheet_name: Sheet name. If None, uses active sheet.

    Returns:
        CellStyle object.

    Raises:
        FileNotFoundError: If file does not exist.
        ValueError: If sheet_name is invalid.
    """
    cell = read_cell(path, coordinate, sheet_name)
    return cell.style


def _get_cell_type(cell: OpenpyxlCell) -> str:
    """Determine cell type code."""
    if cell.value is None:
        return "n"
    if isinstance(cell.value, bool):
        return "b"
    if isinstance(cell.value, datetime):
        return "d"
    if isinstance(cell.value, (int, float)):
        return "n"
    if cell.data_type == "e":
        return "e"
    return "s"


def _get_cell_value(cell: OpenpyxlCell) -> Any:
    """Get cell value, converting datetime to ISO string."""
    if isinstance(cell.value, datetime):
        return cell.value.isoformat()
    return cell.value


def _get_formula(cell: OpenpyxlCell) -> Optional[str]:
    """Extract formula from cell if present."""
    if cell.data_type == "f":
        return cell.value if isinstance(cell.value, str) else None
    if hasattr(cell, 'value') and isinstance(cell.value, str) and cell.value.startswith('='):
        return cell.value
    return None


def read_sheet(
    path: Union[str, Path],
    sheet_name: Optional[str] = None,
) -> Sheet:
    """
    Read entire sheet from an Excel file.

    Args:
        path: Path to the Excel file.
        sheet_name: Sheet name. If None, uses active sheet.

    Returns:
        Sheet object with all data.

    Raises:
        FileNotFoundError: If file does not exist.
        ValueError: If sheet_name is invalid.
    """
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"File not found: {path}")

    wb = load_workbook(path, data_only=False)

    if sheet_name:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet not found: {sheet_name}")
        ws = wb[sheet_name]
    else:
        ws = wb.active

    # Get sheet ID
    sheet_id = 1
    for idx, name in enumerate(wb.sheetnames, 1):
        if name == ws.title:
            sheet_id = idx
            break

    # Build dimensions
    dimensions = SheetDimensions(
        used_range=ws.dimensions if ws.dimensions else None,
        default_row_height=ws.sheet_format.defaultRowHeight or 15.0,
        default_col_width=ws.sheet_format.defaultColWidth or 8.43,
    )

    # Column dimensions
    for col_letter, col_dim in ws.column_dimensions.items():
        if col_dim.width is not None or col_dim.hidden:
            dimensions.columns[col_letter] = ColumnDimension(
                width=col_dim.width or 8.43,
                hidden=col_dim.hidden or False,
                style=col_dim.style if hasattr(col_dim, 'style') else None,
            )

    # Row dimensions
    for row_num, row_dim in ws.row_dimensions.items():
        if row_dim.height is not None or row_dim.hidden:
            dimensions.rows[str(row_num)] = RowDimension(
                height=row_dim.height or 15.0,
                hidden=row_dim.hidden or False,
                style=row_dim.style if hasattr(row_dim, 'style') else None,
            )

    # Merged cells
    merged_cells = [str(merged) for merged in ws.merged_cells.ranges]

    # Cells with data
    cells = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None or cell.data_type == "f":
                coord = cell.coordinate
                formula = _get_formula(cell)
                
                cells[coord] = CellData(
                    v=_get_cell_value(cell) if formula is None else cell.value,
                    t=_get_cell_type(cell),
                    s=cell.style_id if hasattr(cell, 'style_id') and cell.style_id else None,
                    f=formula,
                )

    return Sheet(
        sheet_id=sheet_id,
        name=ws.title,
        dimensions=dimensions,
        merged_cells=merged_cells,
        cells=cells,
    )
