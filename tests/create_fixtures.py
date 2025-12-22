"""Generate test fixture Excel files."""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, Alignment


def create_simple_xlsx(path: Path):
    """Create simple.xlsx with plain text in A1."""
    wb = Workbook()
    sheet = wb.active
    sheet["A1"] = "Hello"
    wb.save(path)


def create_styled_xlsx(path: Path):
    """Create styled.xlsx with formatted A1."""
    wb = Workbook()
    sheet = wb.active
    sheet["A1"] = "Styled"

    sheet["A1"].font = Font(
        name="Arial",
        size=14,
        bold=True,
        italic=False,
        color="FF0000",
    )
    sheet["A1"].fill = PatternFill(
        patternType="solid",
        fgColor="FFFF00",
    )
    sheet["A1"].border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )
    sheet["A1"].alignment = Alignment(
        horizontal="center",
        vertical="center",
    )

    wb.save(path)


def create_empty_xlsx(path: Path):
    """Create empty.xlsx with empty A1."""
    wb = Workbook()
    sheet = wb.active
    # A1 is empty by default
    wb.save(path)


def create_formula_xlsx(path: Path):
    """Create formula.xlsx with formula in A1."""
    wb = Workbook()
    sheet = wb.active
    sheet["A1"] = 10
    sheet["A2"] = 20
    sheet["A3"] = "=A1+A2"
    wb.save(path)


def create_complex_xlsx(path: Path):
    """Create complex.xlsx with various data types and merged cells."""
    wb = Workbook()
    sheet = wb.active
    sheet.title = "Data"
    
    # Header row (merged)
    sheet.merge_cells("A1:D1")
    sheet["A1"] = "Sales Report Q1 2025"
    sheet["A1"].font = Font(bold=True, size=14)
    sheet["A1"].alignment = Alignment(horizontal="center")
    
    # Column headers
    sheet["A2"] = "Product"
    sheet["B2"] = "Quantity"
    sheet["C2"] = "Price"
    sheet["D2"] = "Total"
    for col in ["A", "B", "C", "D"]:
        sheet[f"{col}2"].font = Font(bold=True)
    
    # Data rows
    sheet["A3"] = "Widget"
    sheet["B3"] = 100
    sheet["C3"] = 9.99
    sheet["D3"] = "=B3*C3"
    
    sheet["A4"] = "Gadget"
    sheet["B4"] = 50
    sheet["C4"] = 19.99
    sheet["D4"] = "=B4*C4"
    
    sheet["A5"] = "Gizmo"
    sheet["B5"] = 75
    sheet["C5"] = 14.99
    sheet["D5"] = "=B5*C5"
    
    # Total row
    sheet["A6"] = "Total"
    sheet["A6"].font = Font(bold=True)
    sheet["D6"] = "=SUM(D3:D5)"
    sheet["D6"].font = Font(bold=True)
    
    # Column widths
    sheet.column_dimensions["A"].width = 15
    sheet.column_dimensions["B"].width = 10
    sheet.column_dimensions["C"].width = 10
    sheet.column_dimensions["D"].width = 12
    
    # Row height
    sheet.row_dimensions[1].height = 25
    
    wb.save(path)


def main():
    fixtures_dir = Path(__file__).parent / "fixtures"
    fixtures_dir.mkdir(exist_ok=True)

    create_simple_xlsx(fixtures_dir / "simple.xlsx")
    create_styled_xlsx(fixtures_dir / "styled.xlsx")
    create_empty_xlsx(fixtures_dir / "empty.xlsx")
    create_formula_xlsx(fixtures_dir / "formula.xlsx")
    create_complex_xlsx(fixtures_dir / "complex.xlsx")

    print(f"Created test fixtures in {fixtures_dir}")


if __name__ == "__main__":
    main()
